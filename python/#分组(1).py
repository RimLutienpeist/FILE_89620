#分组

import random
import tkinter
import openpyxl

from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkinter.simpledialog import askstring
from tkinter import filedialog
from openpyxl.styles import PatternFill
from enum import Enum

class StudentBasic:
    def __init__(self, name, classNumber, gender, major) -> None:
        self.name = name
        self.classNumber = classNumber
        self.gender = gender
        self.major = major
        
def defaultResponse():
    msg = tkinter.messagebox.showinfo('Default Response', 'Default Response')

def get_file(enterFileName, enterWorksheets):
    root = Tk()
    file_path = filedialog.askopenfilename(filetypes=[('Excel Workbooks', '*.xlsx')])
    print("所选文件名为：", str(file_path))
    root.destroy()
    enterFileName.config(text= file_path)
    
    workbook = openpyxl.load_workbook(file_path)
    enterWorksheets['value'] = workbook.sheetnames
    print("工作表列表：", workbook.sheetnames)
    workbook.close()
    
def divide():
    workbook = openpyxl.load_workbook(school_EnterFileName.cget('text'))
    worksheet = workbook.get_sheet_by_name(school_EnterWorksheet.get())
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    otherMajorBoys = [] #2
    mixedBoys = [] #4
    otherMajorGirls = [] #1
    mixedGirls = [] #3
    
    for i in range(1, worksheet.max_row+1):
        stu = StudentBasic(worksheet.cell(row=i, column=3).value, worksheet.cell(row=i, column=2).value, worksheet.cell(row=i, column=4).value, worksheet.cell(row=i, column=5).value)
        if stu.gender == '男':
            if stu.major == '混合班':
                mixedBoys.append(stu)
            else:
                otherMajorBoys.append(stu)
        else:
            if stu.major == '混合班':
                mixedGirls.append(stu)
            else:
                otherMajorGirls.append(stu)
    
    i = 1
    j = 1
    while len(otherMajorGirls) > 0:
        mg = otherMajorGirls
        random.shuffle(mg)
        k = random.randint(0, len(otherMajorGirls)-1)
        print(k, mg[k].name)
        ws.cell(row=i, column=j).value = mg[k].name
        ws.cell(row=i, column=j).fill = PatternFill('solid', fgColor='00BFFF')
        otherMajorGirls.pop(k)
        j += 1
        if j ==11:
            j = 1
            i += 1
    while len(otherMajorBoys) > 0:
        mg = otherMajorBoys
        random.shuffle(mg)
        k = random.randint(0, len(otherMajorBoys)-1)
        print(k, mg[k].name)
        ws.cell(row=i, column=j).value = mg[k].name
        ws.cell(row=i, column=j).fill = PatternFill('solid', fgColor='FF69B4')
        otherMajorBoys.pop(k)
        j += 1
        if j ==11:
            j = 1
            i += 1
    while len(mixedGirls) > 0:
        mg = mixedGirls
        random.shuffle(mg)
        k = random.randint(0, len(mixedGirls)-1)
        print(k, mg[k].name)
        ws.cell(row=i, column=j).value = mg[k].name
        ws.cell(row=i, column=j).fill = PatternFill('solid', fgColor='7FFF00')
        mixedGirls.pop(k)
        j += 1
        if j ==11:
            j = 1
            i += 1
    while len(mixedBoys) > 0:
        mg = mixedBoys
        random.shuffle(mg)
        k = random.randint(0, len(mixedBoys)-1)
        print(k, mg[k].name)
        ws.cell(row=i, column=j).value = mg[k].name
        ws.cell(row=i, column=j).fill = PatternFill('solid', fgColor='FFFFFF')
        mixedBoys.pop(k)
        j += 1
        if j ==11:
            j = 1
            i += 1
    
    wb.save('分组结果.xlsx')

window = tkinter.Tk()

# 窗体
window.title('自动表格处理')
window.geometry('500x600')
window.resizable(False, False)

# 功能区栏
taskbar = ttk.Notebook(window)

school = ttk.Frame(taskbar, relief='ridge', borderwidth=1)
school.pack(fill=tkinter.BOTH, expand=True)
taskbar.add(school, text='分组', padding=10)

school_main = tkinter.LabelFrame(school, text= '主要文件', height=400)
school_main.pack(padx=10, pady=10, ipadx=5, ipady=5, fill='x', expand=True)
school_Label_EnterFileName = ttk.Label(school_main, text='请指定Excel文件路径').grid(row=0, column=0)
school_EnterFileName = ttk.Label(school_main)
school_EnterFileName.grid(row=0, column=1)

chooseFile = tkinter.Button(school_main, text='浏览', command=lambda: get_file(school_EnterFileName, school_EnterWorksheet))
chooseFile.grid(row=0, column=3)

school_Label_EnterWorksheet = ttk.Label(school_main, text='请选择目标工作表名')
school_Label_EnterWorksheet.grid(row=1, column=0)
school_EnterWorksheet = ttk.Combobox(school_main, values="")
school_EnterWorksheet.grid(row=1, column=1)
school_EnterWorksheet.insert(0, 'Sheet1')

execute = tkinter.Button(school, text='一键匹配', command=divide)
execute.pack(padx=10, pady=10)

taskbar.pack(side=TOP, fill=BOTH, expand=True)

window.mainloop()